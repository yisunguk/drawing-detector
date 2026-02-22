"""
PlantSync AI - Plant Drawing Revision & Discipline Collaboration API

- POST   /projects                                    : Create project
- GET    /projects                                    : List projects
- GET    /projects/{id}                               : Project detail
- PATCH  /projects/{id}                               : Rename project
- DELETE /projects/{id}                               : Delete project
- POST   /projects/{id}/upload                        : Upload drawing PDF → DI → Title Block
- POST   /projects/{id}/bulk-upload                   : Bulk upload multiple PDFs
- PUT    /projects/{id}/drawings/{did}/title-block     : Confirm/edit Title Block
- POST   /projects/{id}/drawings/{did}/register        : Register staged drawing
- GET    /projects/{id}/staged                         : List staged drawings
- GET    /projects/{id}/drawings/{did}/pdf-url         : Get PDF SAS URL
- POST   /projects/{id}/drawings/{did}/diff-urls       : Get SAS URLs for two revisions (diff)
- POST   /projects/{id}/drawings/{did}/markups         : Create markup (pin)
- GET    /projects/{id}/drawings/{did}/markups          : List markups (filter: page, discipline)
- PATCH  /projects/{id}/drawings/{did}/markups/{mid}   : Update markup status
- POST   /projects/{id}/drawings/{did}/markups/{mid}/replies : Add reply
- POST   /projects/{id}/drawings/{did}/nearby-text     : Find text near pin coordinate
- POST   /projects/{id}/drawings/{did}/related-search  : Search related markups/documents
- PUT    /projects/{id}/drawings/{did}/review           : Update discipline review status
- POST   /projects/{id}/drawings/{did}/approve          : EM final approval
- GET    /projects/{id}/dashboard                       : Dashboard stats
- GET    /projects/{id}/export-excel                   : Export drawing register as Excel
- GET    /projects/{id}/activity                       : Activity timeline
- GET    /projects/{id}/drawings/{did}/review-gate     : Review gate check
- POST   /projects/{id}/drawings/{did}/export-markup-pdf : Export PDF with markup overlays
"""

import io
import json
import logging
import math
import re
import uuid
from datetime import datetime, timezone
from typing import List, Optional

from fastapi import APIRouter, HTTPException, Header, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.core.config import settings
from app.core.firebase_admin import verify_id_token
from app.services.plantsync_firestore import (
    fs_create_project,
    fs_get_project,
    fs_update_project,
    fs_delete_project,
    fs_list_projects,
    fs_get_drawing,
    fs_add_drawing,
    fs_update_drawing,
    fs_delete_drawing,
    fs_list_drawings,
    fs_find_drawing_by_number,
    fs_add_markup,
    fs_get_markup,
    fs_update_markup,
    fs_delete_markup,
    fs_list_markups,
    fs_list_markups_all,
    fs_delete_markups_by_drawing,
    fs_bulk_update_markups,
    fs_add_request,
    fs_get_request,
    fs_update_request,
    fs_delete_request,
    fs_list_requests,
    fs_count_transmittals,
    fs_add_activity,
    fs_list_activities,
    fs_add_shared_ref,
    fs_delete_shared_ref,
    fs_resolve_project,
)

logger = logging.getLogger(__name__)
router = APIRouter()


# ── Pydantic Models ──

class CreateProjectRequest(BaseModel):
    project_name: str
    project_code: Optional[str] = None
    description: Optional[str] = None


class ConfirmTitleBlockRequest(BaseModel):
    drawing_number: Optional[str] = None
    title: Optional[str] = None
    revision: Optional[str] = None
    discipline: Optional[str] = None
    # EPC 관리 필드
    vendor_drawing_number: Optional[str] = None
    issue_purpose: Optional[str] = None  # IFA / IFI / IFC / As-Built
    issue_date: Optional[str] = None
    receive_date: Optional[str] = None
    vendor_name: Optional[str] = None
    reviewer_name: Optional[str] = None
    has_dwg: Optional[bool] = None
    related_drawings: Optional[List[str]] = None
    change_log: Optional[str] = None
    remarks: Optional[str] = None


class CreateMarkupRequest(BaseModel):
    page: int
    x: float
    y: float
    discipline: str
    comment: str
    severity: Optional[str] = "normal"
    author_name: Optional[str] = None
    request_id: Optional[str] = None  # link to review request


class UpdateMarkupRequest(BaseModel):
    status: Optional[str] = None
    comment: Optional[str] = None


class AddMarkupReplyRequest(BaseModel):
    author_name: Optional[str] = None
    content: str


class UpdateReviewStatusRequest(BaseModel):
    discipline: str
    status: str  # "not_started" | "in_progress" | "completed" | "rejected"
    reviewer_name: Optional[str] = None
    comment: Optional[str] = None


class FinalApprovalRequest(BaseModel):
    decision: str  # "approved" | "rejected" | "conditionally_approved"
    approver_name: Optional[str] = None
    comment: Optional[str] = None


class CreateReviewRequestModel(BaseModel):
    drawing_id: str
    to_name: str  # backward compat - maps to lead_reviewer
    discipline: str
    title: str
    message: str = ""
    priority: Optional[str] = "normal"
    # New EPC workflow fields
    due_date: Optional[str] = None
    lead_reviewer: Optional[str] = None
    squad_reviewers: Optional[List[str]] = None
    return_code: Optional[str] = None  # code_1 | code_2 | code_3 | code_4
    transmittal_no: Optional[str] = None


class ReplyReviewRequestModel(BaseModel):
    content: str
    author_name: Optional[str] = None


class UpdateReviewRequestStatusModel(BaseModel):
    status: str  # intake | assigned | markup_in_progress | markup_done | consolidation | return_decided | transmitted | rejected


class IntakeDecisionRequest(BaseModel):
    drawing_id: str
    decision: str  # "accepted" | "rejected_intake"
    vdrl_match: Optional[bool] = None
    comment: Optional[str] = ""


class AssignReviewersRequest(BaseModel):
    lead_reviewer: str
    squad_reviewers: Optional[List[str]] = None
    due_date: Optional[str] = None


class ConsolidateRequest(BaseModel):
    confirmed_markup_ids: Optional[List[str]] = None
    comment: Optional[str] = ""


class ReturnCodeRequest(BaseModel):
    return_code: str  # code_1 | code_2 | code_3 | code_4
    comment: Optional[str] = ""


class TransmittalRequest(BaseModel):
    comment: Optional[str] = ""


# ── Auth Helper ──

def _get_username(authorization: Optional[str]) -> str:
    if not authorization or not authorization.startswith('Bearer '):
        raise HTTPException(status_code=401, detail="Missing or invalid Authorization header")

    token = authorization.replace('Bearer ', '')
    try:
        decoded = verify_id_token(token)
        name = decoded.get('name')
        if not name:
            uid = decoded.get('uid')
            if uid:
                try:
                    from firebase_admin import firestore
                    db = firestore.client()
                    user_doc = db.collection('users').document(uid).get()
                    if user_doc.exists:
                        user_data = user_doc.to_dict()
                        name = user_data.get('name') or user_data.get('displayName')
                except Exception:
                    pass
        return name or decoded.get('email', '').split('@')[0] or 'unknown'
    except Exception as e:
        raise HTTPException(status_code=401, detail=f"Authentication failed: {e}")


# ── Blob Helper (PDF only) ──

def _get_container():
    from app.services.blob_storage import get_container_client
    return get_container_client()


def _load_json(container, blob_path: str) -> Optional[dict]:
    """Load JSON from blob — kept only for DI result files."""
    try:
        blob = container.get_blob_client(blob_path)
        data = blob.download_blob().readall()
        return json.loads(data.decode('utf-8'))
    except Exception as e:
        logger.debug(f"Failed to load JSON: {blob_path} - {e}")
        return None


def _save_json(container, blob_path: str, obj):
    """Save JSON to blob — kept only for DI result files."""
    data = json.dumps(obj, ensure_ascii=False, indent=2).encode('utf-8')
    container.upload_blob(name=blob_path, data=data, overwrite=True)


DISCIPLINES = [
    "process", "mechanical", "piping", "electrical", "instrument", "civil"
]


def _empty_review_status():
    return {d: {"status": "not_started"} for d in DISCIPLINES}


# ── Title Block Extraction ──

def _extract_title_block(di_pages: list) -> dict:
    """Extract drawing metadata from DI result (all pages, title block focus on first page)."""
    result = {
        "drawing_number": "",
        "title": "",
        "revision": "",
        "discipline": "",
        # EPC fields — auto-extracted
        "issue_purpose": "",
        "issue_date": "",
        "vendor_drawing_number": "",
        "vendor_name": "",
    }
    if not di_pages:
        return result

    page = di_pages[0]
    content = page.get("content", "")

    # Combine all pages content for broader search
    all_content = "\n".join(p.get("content", "") for p in di_pages)

    # Try DI metadata fields first
    if page.get("도면번호(DWG. NO.)"):
        result["drawing_number"] = page["도면번호(DWG. NO.)"]
    if page.get("도면명(TITLE)"):
        result["title"] = page["도면명(TITLE)"]

    # Pattern matching from content
    lines = content.split('\n')
    all_lines = all_content.split('\n')

    for line in lines:
        line_upper = line.strip().upper()

        # Drawing number patterns
        if not result["drawing_number"]:
            dwg_match = re.search(
                r'(?:DWG\.?\s*(?:NO\.?|NUMBER)?|DRAWING\s*(?:NO\.?|NUMBER)?)\s*[:\-]?\s*([A-Z0-9][\w\-\.]+)',
                line_upper
            )
            if dwg_match:
                result["drawing_number"] = dwg_match.group(1).strip()

            # Generic drawing number pattern (e.g. 147600-012-SZ-0002, 10-24000-OM-171-200)
            if not result["drawing_number"]:
                generic = re.search(r'(\d{2,}-\d{3,}-[A-Z]{2,}-[\d\-]+)', line_upper)
                if generic:
                    result["drawing_number"] = generic.group(1)

        # Revision patterns
        if not result["revision"]:
            rev_match = re.search(
                r'(?:REV\.?\s*(?:NO\.?)?|REVISION)\s*[:\-]?\s*([A-Z0-9]{1,3})',
                line_upper
            )
            if rev_match:
                result["revision"] = rev_match.group(1).strip()

        # Title patterns
        if not result["title"]:
            title_match = re.search(
                r'(?:TITLE|DESCRIPTION)\s*[:\-]?\s*(.{5,})',
                line_upper
            )
            if title_match:
                result["title"] = title_match.group(1).strip()

    # If no title found from patterns, use the DI page content snippet
    if not result["title"] and len(lines) > 2:
        longest = max(lines[:20], key=len).strip() if lines else ""
        if len(longest) > 5:
            result["title"] = longest[:200]

    # ── Discipline: drawing number code → discipline mapping ──
    dwg_lower = result["drawing_number"].lower()
    content_lower = content.lower()

    # Discipline code map (common EPC 2-letter codes in drawing numbers)
    DISC_CODE_MAP = {
        "process":     ['-pf-', '-pd-', '-ph-', '-sz-'],
        "mechanical":  ['-me-', '-mc-', '-om-', '-mh-', '-mv-', '-ma-'],
        "piping":      ['-pp-', '-pi-', '-pl-', '-pa-', '-pb-', '-ps-'],
        "electrical":  ['-el-', '-ee-', '-ea-', '-ec-', '-ed-'],
        "instrument":  ['-in-', '-ic-', '-ia-', '-id-'],
        "civil":       ['-cv-', '-cs-', '-ci-', '-ca-', '-st-'],
    }

    if any(k in dwg_lower or k in content_lower for k in ['pid', 'p&id', 'process']):
        result["discipline"] = "process"
    else:
        for disc, codes in DISC_CODE_MAP.items():
            if any(c in dwg_lower for c in codes):
                result["discipline"] = disc
                break

    # ── Issue Purpose extraction ──
    # Look for IFA / IFI / IFC / IFD / IFR / As-Built / AFC / AFD patterns
    PURPOSE_PATTERNS = [
        (r'\bIFA\b', "IFA"),
        (r'\bIFI\b', "IFI"),
        (r'\bIFC\b', "IFC"),
        (r'\bIFD\b', "IFD"),
        (r'\bIFR\b', "IFR"),
        (r'\bAFC\b', "AFC"),
        (r'\bAFD\b', "AFD"),
        (r'\bAS[\s\-]?BUILT\b', "As-Built"),
        (r'ISSUED?\s+FOR\s+APPROVAL', "IFA"),
        (r'ISSUED?\s+FOR\s+INFORMATION', "IFI"),
        (r'ISSUED?\s+FOR\s+CONSTRUCTION', "IFC"),
        (r'ISSUED?\s+FOR\s+DESIGN', "IFD"),
        (r'ISSUED?\s+FOR\s+REVIEW', "IFR"),
        (r'APPROVED?\s+FOR\s+CONSTRUCTION', "AFC"),
        (r'APPROVED?\s+FOR\s+DESIGN', "AFD"),
        (r'FOR\s+APPROVAL', "IFA"),
        (r'FOR\s+CONSTRUCTION', "IFC"),
        (r'FOR\s+INFORMATION', "IFI"),
    ]
    all_upper = all_content.upper()
    for pattern, purpose in PURPOSE_PATTERNS:
        if re.search(pattern, all_upper):
            result["issue_purpose"] = purpose
            break

    # ── Issue Date extraction ──
    # Search near keywords like "DATE", "ISSUE DATE", "ISSUED"
    DATE_PATTERN = r'(\d{4}[\.\-/]\d{1,2}[\.\-/]\d{1,2}|\d{1,2}[\.\-/]\d{1,2}[\.\-/]\d{4}|\d{1,2}\s+(?:JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)[A-Z]*\.?\s+\d{4}|\d{4}\s*년\s*\d{1,2}\s*월\s*\d{1,2}\s*일)'
    for line in all_lines:
        line_upper = line.strip().upper()
        # Prioritize lines containing date-related keywords
        if any(k in line_upper for k in ['ISSUE DATE', 'ISSUED', 'DATE OF ISSUE', '발행일']):
            date_match = re.search(DATE_PATTERN, line, re.IGNORECASE)
            if date_match:
                result["issue_date"] = _normalize_date(date_match.group(0))
                break
    # Fallback: search for DATE keyword + date on same line
    if not result["issue_date"]:
        for line in all_lines:
            line_upper = line.strip().upper()
            if 'DATE' in line_upper and not any(k in line_upper for k in ['RECEIVE', 'RETURN', 'DUE', 'EXPIR']):
                date_match = re.search(DATE_PATTERN, line, re.IGNORECASE)
                if date_match:
                    result["issue_date"] = _normalize_date(date_match.group(0))
                    break

    # ── Vendor Drawing Number ──
    VDN_RE = re.compile(
        r'(?:VENDOR|SUPPLIER)\s+(?:DWG|DRAWING|DOC)\s*(?:NO\.?|NUMBER|#)?\s*[:\-]?\s*([A-Z0-9][\w\-\.]{3,})'
    )
    for line in all_lines:
        vdn_match = VDN_RE.search(line.strip().upper())
        if vdn_match:
            result["vendor_drawing_number"] = vdn_match.group(1).strip()
            break

    # ── Vendor Name ──
    VENDOR_KEYWORDS = ['VENDOR', 'SUPPLIER', 'MANUFACTURER', 'PREPARED BY', 'MADE BY', 'DESIGNED BY', 'CONTRACTOR']
    for line in all_lines:
        line_upper = line.strip().upper()
        for kw in VENDOR_KEYWORDS:
            if kw in line_upper and 'DWG' not in line_upper and 'DRAWING' not in line_upper and 'DOC' not in line_upper:
                # Extract value after the keyword
                vn_match = re.search(
                    rf'{kw}\s*[:\-]?\s*(.{{2,50}})',
                    line_upper
                )
                if vn_match:
                    name = vn_match.group(1).strip().strip(':- ')
                    # Filter out noise — should look like a company name
                    if name and len(name) >= 2 and not name.startswith(('NO', 'NUMBER', '#', 'DATE')):
                        result["vendor_name"] = line.strip()[line.upper().strip().index(name):line.upper().strip().index(name)+len(name)]
                        break
        if result["vendor_name"]:
            break

    return result


def _normalize_date(raw: str) -> str:
    """Normalize various date formats to YYYY-MM-DD."""
    raw = raw.strip()

    # 2024.05.15 / 2024-05-15 / 2024/05/15
    m = re.match(r'(\d{4})[\.\-/](\d{1,2})[\.\-/](\d{1,2})', raw)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"

    # 15.05.2024 / 15-05-2024 / 15/05/2024
    m = re.match(r'(\d{1,2})[\.\-/](\d{1,2})[\.\-/](\d{4})', raw)
    if m:
        day, month, year = int(m.group(1)), int(m.group(2)), m.group(3)
        if day > 12:  # clearly DD/MM/YYYY
            return f"{year}-{month:02d}-{day:02d}"
        return f"{year}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"

    # 15 MAY 2024 / 15 May 2024
    MONTHS = {'JAN': '01', 'FEB': '02', 'MAR': '03', 'APR': '04', 'MAY': '05', 'JUN': '06',
              'JUL': '07', 'AUG': '08', 'SEP': '09', 'OCT': '10', 'NOV': '11', 'DEC': '12'}
    m = re.match(r'(\d{1,2})\s+([A-Za-z]+)\.?\s+(\d{4})', raw)
    if m:
        month_str = m.group(2).upper()[:3]
        if month_str in MONTHS:
            return f"{m.group(3)}-{MONTHS[month_str]}-{int(m.group(1)):02d}"

    # 2024년 5월 15일
    m = re.match(r'(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일', raw)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"

    return raw


# ── API Endpoints ──

@router.get("/users")
async def list_users(authorization: Optional[str] = Header(None)):
    """Return list of registered users from Firestore."""
    _get_username(authorization)  # auth check
    try:
        from firebase_admin import firestore
        db = firestore.client()
        docs = db.collection('users').stream()
        users = []
        for doc in docs:
            d = doc.to_dict()
            name = d.get('name') or d.get('displayName') or ''
            email = d.get('email') or ''
            if name or email:
                users.append({"uid": doc.id, "name": name, "email": email})
        users.sort(key=lambda u: u['name'] or u['email'])
        return {"users": users}
    except Exception as e:
        logger.error(f"list_users error: {e}")
        return {"users": []}


@router.post("/projects")
async def create_project(
    req: CreateProjectRequest,
    authorization: Optional[str] = Header(None)
):
    username = _get_username(authorization)

    project_id = str(uuid.uuid4())[:8]
    now = datetime.now(timezone.utc).isoformat()

    project_data = {
        "project_id": project_id,
        "project_name": req.project_name,
        "project_code": req.project_code or "",
        "description": req.description or "",
        "owner": username,
        "members": [],
        "member_names": [],  # denormalized for Firestore array-contains query
        "created_at": now,
        "updated_at": now,
    }

    fs_create_project(project_id, project_data)
    print(f"[PlantSync] Project created: {project_id} by {username}", flush=True)

    # Return response with drawing_count for frontend compatibility
    response_data = {**project_data, "drawings": []}
    return {"status": "success", "project": response_data}


@router.get("/projects")
async def list_projects(authorization: Optional[str] = Header(None)):
    username = _get_username(authorization)

    projects = fs_list_projects(username)

    summaries = []
    for p in projects:
        pid = p["project_id"]
        drawing_count = len(fs_list_drawings(pid))
        is_shared = p.get("owner") != username
        summaries.append({
            "project_id": pid,
            "project_name": p["project_name"],
            "project_code": p.get("project_code", ""),
            "description": p.get("description", ""),
            "drawing_count": drawing_count,
            "created_at": p.get("created_at", ""),
            "updated_at": p.get("updated_at", ""),
            "is_shared": is_shared,
            "owner_name": p.get("owner", username),
        })

    return {"status": "success", "projects": summaries}


@router.get("/projects/{project_id}")
async def get_project(project_id: str, authorization: Optional[str] = Header(None)):
    username = _get_username(authorization)
    project = fs_resolve_project(username, project_id)
    drawings = fs_list_drawings(project_id)
    project["drawings"] = drawings
    return {"status": "success", "project": project}


class RenameProjectRequest(BaseModel):
    project_name: str


@router.patch("/projects/{project_id}")
async def rename_project(project_id: str, req: RenameProjectRequest, authorization: Optional[str] = Header(None)):
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    now = datetime.now(timezone.utc).isoformat()
    fs_update_project(project_id, {
        "project_name": req.project_name.strip(),
        "updated_at": now,
    })
    print(f"[PlantSync] Project renamed: {project_id} -> {req.project_name}", flush=True)

    project = fs_get_project(project_id)
    project["drawings"] = fs_list_drawings(project_id)
    return {"status": "success", "project": project}


@router.delete("/projects/{project_id}")
async def delete_project(project_id: str, authorization: Optional[str] = Header(None)):
    username = _get_username(authorization)
    project = fs_resolve_project(username, project_id)

    # Only owner can delete
    if project.get("owner") != username:
        raise HTTPException(status_code=403, detail="Only the project owner can delete a project")

    # Clean up shared references for all members
    for m in project.get("members", []):
        fs_delete_shared_ref(m.get("name", ""), project_id)

    # Delete blob files for all drawings
    container = _get_container()
    drawings = fs_list_drawings(project_id)
    blob_count = 0
    for d in drawings:
        for rev in d.get("revisions", []):
            bp = rev.get("blob_path", "")
            if bp:
                try:
                    container.delete_blob(bp)
                    blob_count += 1
                except Exception:
                    pass
                di_path = bp.rsplit('/', 1)[0] + "/di_result.json"
                try:
                    container.delete_blob(di_path)
                    blob_count += 1
                except Exception:
                    pass

    # Delete Firestore project + subcollections
    fs_delete_project(project_id)
    print(f"[PlantSync] Project deleted: {project_id}, {blob_count} blobs removed", flush=True)

    return {"status": "success", "deleted": project_id}


class AddMemberRequest(BaseModel):
    member_uid: str
    member_name: str
    member_email: Optional[str] = ""


@router.post("/projects/{project_id}/members")
async def add_member(
    project_id: str,
    req: AddMemberRequest,
    authorization: Optional[str] = Header(None)
):
    """Add a member to the project. Creates a shared reference."""
    username = _get_username(authorization)
    project = fs_resolve_project(username, project_id)

    members = project.get("members", [])
    member_names = project.get("member_names", [])

    # Prevent duplicates
    if any(m.get("name") == req.member_name for m in members):
        return {"status": "success", "message": "Already a member", "members": members}

    members.append({"uid": req.member_uid, "name": req.member_name, "email": req.member_email or ""})
    member_names.append(req.member_name)

    now = datetime.now(timezone.utc).isoformat()
    fs_update_project(project_id, {
        "members": members,
        "member_names": member_names,
        "updated_at": now,
    })

    # Create shared reference
    fs_add_shared_ref(req.member_name, project_id, {
        "owner": project.get("owner"),
        "project_name": project.get("project_name", ""),
    })

    fs_add_activity(project_id, "member_added", {
        "member_name": req.member_name,
    }, actor=username)

    print(f"[PlantSync] Member added: {req.member_name} to {project_id}", flush=True)
    return {"status": "success", "members": members}


@router.delete("/projects/{project_id}/members/{member_name}")
async def remove_member(
    project_id: str,
    member_name: str,
    authorization: Optional[str] = Header(None)
):
    """Remove a member from the project. Owner only."""
    username = _get_username(authorization)
    project = fs_resolve_project(username, project_id)

    # Only owner can remove members
    if project.get("owner") != username:
        raise HTTPException(status_code=403, detail="Only the project owner can remove members")

    members = project.get("members", [])
    new_members = [m for m in members if m.get("name") != member_name]
    new_member_names = [m.get("name") for m in new_members]

    now = datetime.now(timezone.utc).isoformat()
    fs_update_project(project_id, {
        "members": new_members,
        "member_names": new_member_names,
        "updated_at": now,
    })

    # Delete shared reference
    fs_delete_shared_ref(member_name, project_id)

    fs_add_activity(project_id, "member_removed", {
        "member_name": member_name,
    }, actor=username)

    print(f"[PlantSync] Member removed: {member_name} from {project_id}", flush=True)
    return {"status": "success", "members": new_members}


@router.get("/projects/{project_id}/members")
async def list_members(
    project_id: str,
    authorization: Optional[str] = Header(None)
):
    """List project members."""
    username = _get_username(authorization)
    project = fs_resolve_project(username, project_id)

    return {
        "status": "success",
        "owner": project.get("owner"),
        "members": project.get("members", []),
    }


@router.post("/projects/{project_id}/upload")
async def upload_drawing(
    project_id: str,
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(None)
):
    username = _get_username(authorization)
    project = fs_resolve_project(username, project_id)
    owner = project["owner"]
    container = _get_container()

    # Read file
    file_content = await file.read()
    if not file_content:
        raise HTTPException(status_code=400, detail="Empty file")

    drawing_id = str(uuid.uuid4())[:8]
    filename = file.filename or f"{drawing_id}.pdf"

    # Upload PDF to temp path first, will move after discipline is determined
    temp_blob_path = f"{owner}/plantsync/{project_id}/temp/{drawing_id}/{filename}"
    container.upload_blob(name=temp_blob_path, data=file_content, overwrite=True)
    print(f"[PlantSync] PDF uploaded (temp): {temp_blob_path}", flush=True)

    # Generate SAS URL for DI analysis
    from app.services.blob_storage import generate_sas_url
    pdf_url = generate_sas_url(temp_blob_path)

    # Run Document Intelligence (all pages, single call)
    di_result = []
    title_block = {"drawing_number": "", "title": "", "revision": "", "discipline": "",
                    "issue_purpose": "", "issue_date": "", "vendor_drawing_number": "", "vendor_name": ""}
    try:
        from app.services.azure_di import azure_di_service
        di_result = azure_di_service.analyze_document_from_url(pdf_url)
        print(f"[PlantSync] DI analysis complete: {len(di_result)} pages", flush=True)

        # Extract title block from first page
        title_block = _extract_title_block(di_result)
        print(f"[PlantSync] Title block extracted: {title_block}", flush=True)
    except Exception as e:
        print(f"[PlantSync] DI analysis failed (non-fatal): {e}", flush=True)

    # Determine discipline folder (fallback to "unknown")
    discipline = title_block.get("discipline") or "unknown"

    # Move PDF from temp to discipline folder
    blob_path = f"{owner}/plantsync/{project_id}/{discipline}/{drawing_id}/{filename}"
    container.upload_blob(name=blob_path, data=file_content, overwrite=True)
    try:
        container.delete_blob(temp_blob_path)
    except Exception:
        pass

    # Save DI result in discipline folder (still as blob — large JSON)
    if di_result:
        di_path = f"{owner}/plantsync/{project_id}/{discipline}/{drawing_id}/di_result.json"
        _save_json(container, di_path, di_result)

    page_count = len(di_result) if di_result else 1

    now = datetime.now(timezone.utc).isoformat()

    # Check if same drawing_number already exists → add as new revision
    existing_drawing = None
    if title_block.get("drawing_number"):
        existing_drawing = fs_find_drawing_by_number(project_id, title_block["drawing_number"])

    if existing_drawing:
        # Add new revision to existing drawing
        rev_entry = {
            "revision_id": drawing_id,
            "revision": title_block.get("revision", ""),
            "blob_path": blob_path,
            "uploaded_at": now,
            "page_count": page_count,
            "filename": filename,
        }
        existing_drawing["revisions"].append(rev_entry)
        existing_drawing["current_revision"] = title_block.get("revision", existing_drawing.get("current_revision", ""))
        existing_drawing["updated_at"] = now

        fs_update_drawing(project_id, existing_drawing["drawing_id"], {
            "revisions": existing_drawing["revisions"],
            "current_revision": existing_drawing["current_revision"],
            "updated_at": now,
        })

        drawing_data = existing_drawing
        is_new_revision = True
        print(f"[PlantSync] New revision added to {existing_drawing['drawing_number']}", flush=True)
    else:
        # Create new drawing entry — auto-fill from DI extraction
        today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        drawing_data = {
            "drawing_id": drawing_id,
            "drawing_number": title_block.get("drawing_number", ""),
            "title": title_block.get("title", filename),
            "discipline": title_block.get("discipline", ""),
            "current_revision": title_block.get("revision", "A"),
            "revisions": [{
                "revision_id": drawing_id,
                "revision": title_block.get("revision", "A"),
                "blob_path": blob_path,
                "uploaded_at": now,
                "page_count": page_count,
                "filename": filename,
            }],
            "review_status": _empty_review_status(),
            "em_approval": {"status": "pending"},
            "staging_status": "staged",
            # EPC 관리 필드 — auto-filled from DI extraction
            "vendor_drawing_number": title_block.get("vendor_drawing_number", ""),
            "issue_purpose": title_block.get("issue_purpose", ""),
            "issue_date": title_block.get("issue_date", ""),
            "receive_date": today_str,  # 접수일 = 업로드 일자
            "vendor_name": title_block.get("vendor_name", ""),
            "reviewer_name": "",
            "has_dwg": False,
            "related_drawings": [],
            "change_log": "",
            "remarks": "",
            "created_at": now,
            "updated_at": now,
        }
        fs_add_drawing(project_id, drawing_id, drawing_data)
        is_new_revision = False

    fs_update_project(project_id, {"updated_at": now})

    fs_add_activity(project_id, "drawing_uploaded", {
        "drawing_id": drawing_id,
        "drawing_number": title_block.get("drawing_number", ""),
        "filename": filename,
        "is_new_revision": is_new_revision,
    }, actor=username)

    # Extract words with confidence & polygon for staging overlay
    title_block_words = []
    di_page_layout = {"width": 0, "height": 0}
    if di_result and len(di_result) > 0:
        first_page = di_result[0]
        di_page_layout = {
            "width": first_page.get("width", 0),
            "height": first_page.get("height", 0),
        }
        for word in first_page.get("words", []):
            title_block_words.append({
                "content": word.get("content", ""),
                "confidence": word.get("confidence", 0),
                "polygon": word.get("polygon", []),
            })

    return {
        "status": "success",
        "drawing": drawing_data,
        "title_block": title_block,
        "is_new_revision": is_new_revision,
        "page_count": page_count,
        "title_block_words": title_block_words,
        "di_page_layout": di_page_layout,
    }


@router.put("/projects/{project_id}/drawings/{drawing_id}/title-block")
async def confirm_title_block(
    project_id: str,
    drawing_id: str,
    req: ConfirmTitleBlockRequest,
    authorization: Optional[str] = Header(None)
):
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    drawing = fs_get_drawing(project_id, drawing_id)
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing not found")

    updates = {}
    if req.drawing_number is not None:
        updates["drawing_number"] = req.drawing_number
    if req.title is not None:
        updates["title"] = req.title
    if req.revision is not None:
        updates["current_revision"] = req.revision
        # Also update latest revision entry
        revisions = drawing.get("revisions", [])
        if revisions:
            revisions[-1]["revision"] = req.revision
            updates["revisions"] = revisions
    if req.discipline is not None:
        updates["discipline"] = req.discipline

    # EPC 관리 필드 처리
    for field in [
        "vendor_drawing_number", "issue_purpose", "issue_date", "receive_date",
        "vendor_name", "reviewer_name", "has_dwg", "change_log", "remarks",
    ]:
        val = getattr(req, field, None)
        if val is not None:
            updates[field] = val
    if req.related_drawings is not None:
        updates["related_drawings"] = req.related_drawings

    now = datetime.now(timezone.utc).isoformat()
    updates["updated_at"] = now
    fs_update_drawing(project_id, drawing_id, updates)
    fs_update_project(project_id, {"updated_at": now})

    drawing = fs_get_drawing(project_id, drawing_id)
    return {"status": "success", "drawing": drawing}


@router.delete("/projects/{project_id}/drawings/{drawing_id}")
async def delete_drawing_endpoint(
    project_id: str,
    drawing_id: str,
    authorization: Optional[str] = Header(None)
):
    username = _get_username(authorization)
    project = fs_resolve_project(username, project_id)
    owner = project["owner"]
    container = _get_container()

    drawing = fs_get_drawing(project_id, drawing_id)
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing not found")

    # Delete blob files for all revisions
    for rev in drawing.get("revisions", []):
        blob_path = rev.get("blob_path", "")
        if blob_path:
            try:
                container.delete_blob(blob_path)
            except Exception:
                pass
            # Also try to delete di_result.json in the same folder
            di_path = blob_path.rsplit('/', 1)[0] + "/di_result.json"
            try:
                container.delete_blob(di_path)
            except Exception:
                pass

    # Remove related markups
    fs_delete_markups_by_drawing(project_id, drawing_id)

    # Remove drawing from Firestore
    fs_delete_drawing(project_id, drawing_id)
    fs_update_project(project_id, {"updated_at": datetime.now(timezone.utc).isoformat()})

    print(f"[PlantSync] Drawing deleted: {drawing_id} ({drawing.get('drawing_number', '')})", flush=True)
    return {"status": "success", "deleted": drawing_id}


@router.get("/projects/{project_id}/drawings/{drawing_id}/pdf-url")
async def get_pdf_url(
    project_id: str,
    drawing_id: str,
    revision_id: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None)
):
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    drawing = fs_get_drawing(project_id, drawing_id)
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing not found")

    # Find the revision
    revisions = drawing.get("revisions", [])
    if not revisions:
        raise HTTPException(status_code=404, detail="No revisions found")

    if revision_id:
        rev = next((r for r in revisions if r["revision_id"] == revision_id), None)
    else:
        rev = revisions[-1]  # Latest

    if not rev:
        raise HTTPException(status_code=404, detail="Revision not found")

    from app.services.blob_storage import generate_sas_url
    pdf_url = generate_sas_url(rev["blob_path"])

    return {
        "status": "success",
        "pdf_url": pdf_url,
        "revision": rev,
    }


# ── Markups ──

@router.post("/projects/{project_id}/drawings/{drawing_id}/markups")
async def create_markup(
    project_id: str,
    drawing_id: str,
    req: CreateMarkupRequest,
    authorization: Optional[str] = Header(None)
):
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    drawing = fs_get_drawing(project_id, drawing_id)
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing not found")

    markup_id = str(uuid.uuid4())[:8]
    now = datetime.now(timezone.utc).isoformat()

    markup = {
        "markup_id": markup_id,
        "drawing_id": drawing_id,
        "request_id": req.request_id or "",
        "page": req.page,
        "x": req.x,
        "y": req.y,
        "discipline": req.discipline,
        "comment": req.comment,
        "severity": req.severity or "normal",
        "status": "open",
        "author_name": req.author_name or username,
        "created_at": now,
        "updated_at": now,
        "replies": [],
    }

    fs_add_markup(project_id, markup_id, markup)

    fs_add_activity(project_id, "markup_created", {
        "markup_id": markup_id,
        "drawing_id": drawing_id,
        "discipline": req.discipline,
        "comment": req.comment[:100],
    }, actor=username)

    return {"status": "success", "markup": markup}


@router.get("/projects/{project_id}/drawings/{drawing_id}/markups")
async def list_markups_endpoint(
    project_id: str,
    drawing_id: str,
    page: Optional[int] = Query(None),
    discipline: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None)
):
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    markups = fs_list_markups(project_id, drawing_id=drawing_id, page=page, discipline=discipline)
    return {"status": "success", "markups": markups}


@router.patch("/projects/{project_id}/drawings/{drawing_id}/markups/{markup_id}")
async def update_markup_endpoint(
    project_id: str,
    drawing_id: str,
    markup_id: str,
    req: UpdateMarkupRequest,
    authorization: Optional[str] = Header(None)
):
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    markup = fs_get_markup(project_id, markup_id)
    if not markup:
        raise HTTPException(status_code=404, detail="Markup not found")

    updates = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if req.status is not None:
        updates["status"] = req.status
    if req.comment is not None:
        updates["comment"] = req.comment

    fs_update_markup(project_id, markup_id, updates)

    if req.status == "resolved":
        fs_add_activity(project_id, "markup_resolved", {
            "markup_id": markup_id, "drawing_id": drawing_id,
        }, actor=username)

    markup.update(updates)
    return {"status": "success", "markup": markup}


@router.post("/projects/{project_id}/drawings/{drawing_id}/markups/{markup_id}/replies")
async def add_markup_reply(
    project_id: str,
    drawing_id: str,
    markup_id: str,
    req: AddMarkupReplyRequest,
    authorization: Optional[str] = Header(None)
):
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    markup = fs_get_markup(project_id, markup_id)
    if not markup:
        raise HTTPException(status_code=404, detail="Markup not found")

    now = datetime.now(timezone.utc).isoformat()
    reply = {
        "reply_id": str(uuid.uuid4())[:8],
        "author_name": req.author_name or username,
        "content": req.content,
        "created_at": now,
    }

    replies = markup.get("replies", [])
    replies.append(reply)

    fs_update_markup(project_id, markup_id, {
        "replies": replies,
        "updated_at": now,
    })

    return {"status": "success", "reply": reply}


# ── Review Workflow ──

@router.put("/projects/{project_id}/drawings/{drawing_id}/review")
async def update_review_status(
    project_id: str,
    drawing_id: str,
    req: UpdateReviewStatusRequest,
    authorization: Optional[str] = Header(None)
):
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    drawing = fs_get_drawing(project_id, drawing_id)
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing not found")

    if req.discipline not in DISCIPLINES:
        raise HTTPException(status_code=400, detail=f"Invalid discipline. Must be one of: {DISCIPLINES}")

    review_status = drawing.get("review_status", _empty_review_status())
    now = datetime.now(timezone.utc).isoformat()

    review_status[req.discipline] = {
        "status": req.status,
        "reviewer_name": req.reviewer_name or username,
        "comment": req.comment or "",
        "updated_at": now,
    }

    fs_update_drawing(project_id, drawing_id, {
        "review_status": review_status,
        "updated_at": now,
    })
    fs_update_project(project_id, {"updated_at": now})

    fs_add_activity(project_id, "review_updated", {
        "drawing_id": drawing_id,
        "discipline": req.discipline,
        "status": req.status,
    }, actor=username)

    return {"status": "success", "review_status": review_status}


@router.post("/projects/{project_id}/drawings/{drawing_id}/approve")
async def final_approval(
    project_id: str,
    drawing_id: str,
    req: FinalApprovalRequest,
    authorization: Optional[str] = Header(None)
):
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    drawing = fs_get_drawing(project_id, drawing_id)
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing not found")

    now = datetime.now(timezone.utc).isoformat()
    em_approval = {
        "status": req.decision,
        "approver_name": req.approver_name or username,
        "comment": req.comment or "",
        "approved_at": now,
    }

    fs_update_drawing(project_id, drawing_id, {
        "em_approval": em_approval,
        "updated_at": now,
    })
    fs_update_project(project_id, {"updated_at": now})

    fs_add_activity(project_id, "approval_decided", {
        "drawing_id": drawing_id,
        "decision": req.decision,
    }, actor=username)

    return {"status": "success", "em_approval": em_approval}


# ── Dashboard ──

@router.get("/projects/{project_id}/dashboard")
async def get_dashboard(
    project_id: str,
    authorization: Optional[str] = Header(None)
):
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    drawings = fs_list_drawings(project_id)
    markups = fs_list_markups_all(project_id)

    # Compute stats
    total_drawings = len(drawings)
    by_discipline = {}
    review_progress = {d: {"not_started": 0, "in_progress": 0, "completed": 0, "rejected": 0} for d in DISCIPLINES}
    approval_stats = {"pending": 0, "approved": 0, "rejected": 0, "conditionally_approved": 0}

    for d in drawings:
        disc = d.get("discipline", "unknown")
        by_discipline[disc] = by_discipline.get(disc, 0) + 1

        # Review progress
        rs = d.get("review_status", {})
        for discipline_key in DISCIPLINES:
            st = rs.get(discipline_key, {}).get("status", "not_started")
            if st in review_progress[discipline_key]:
                review_progress[discipline_key][st] += 1

        # Approval stats
        ea = d.get("em_approval", {}).get("status", "pending")
        if ea in approval_stats:
            approval_stats[ea] += 1

    # Markup stats
    total_markups = len(markups)
    open_markups = sum(1 for m in markups if m.get("status") == "open")
    resolved_markups = sum(1 for m in markups if m.get("status") == "resolved")
    markups_by_discipline = {}
    for m in markups:
        md = m.get("discipline", "unknown")
        markups_by_discipline[md] = markups_by_discipline.get(md, 0) + 1

    # Request stats
    requests = fs_list_requests(project_id)
    total_requests = len(requests)
    pending_requests = sum(1 for r in requests if r.get("status") in ("intake", "assigned", "markup_in_progress", "markup_done", "consolidation", "requested", "feedback"))

    return {
        "status": "success",
        "dashboard": {
            "total_drawings": total_drawings,
            "drawings_by_discipline": by_discipline,
            "review_progress": review_progress,
            "approval_stats": approval_stats,
            "total_markups": total_markups,
            "open_markups": open_markups,
            "resolved_markups": resolved_markups,
            "markups_by_discipline": markups_by_discipline,
            "total_requests": total_requests,
            "pending_requests": pending_requests,
        }
    }


# ── Review Requests (Collaboration) ──

@router.post("/projects/{project_id}/requests")
async def create_review_request(
    project_id: str,
    req: CreateReviewRequestModel,
    authorization: Optional[str] = Header(None)
):
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    # Verify drawing exists
    drawing = fs_get_drawing(project_id, req.drawing_id)
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing not found")

    now = datetime.now(timezone.utc).isoformat()
    request_id = str(uuid.uuid4())[:8]

    review_request = {
        "request_id": request_id,
        "drawing_id": req.drawing_id,
        "drawing_number": drawing.get("drawing_number", ""),
        "from_name": username,
        "to_name": req.to_name,
        "lead_reviewer": req.lead_reviewer or req.to_name,
        "squad_reviewers": req.squad_reviewers or [],
        "discipline": req.discipline,
        "title": req.title,
        "message": req.message,
        "priority": req.priority or "normal",
        "status": "intake",  # New initial status
        "due_date": req.due_date or "",
        "return_code": req.return_code or "",
        "transmittal_no": req.transmittal_no or "",
        "reviewer_statuses": {},
        "created_at": now,
        "updated_at": now,
        "replies": [],
    }

    fs_add_request(project_id, request_id, review_request)

    fs_add_activity(project_id, "request_created", {
        "request_id": request_id,
        "drawing_id": req.drawing_id,
        "to_name": req.to_name,
        "title": req.title,
    }, actor=username)

    print(f"[PlantSync] Review request created: {request_id} ({username} → {req.to_name})", flush=True)
    return {"status": "success", "request": review_request}


@router.get("/projects/{project_id}/requests")
async def list_review_requests(
    project_id: str,
    drawing_id: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None)
):
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    requests = fs_list_requests(project_id, drawing_id=drawing_id, status=status)

    # Enrich with linked markup counts
    all_markups = fs_list_markups_all(project_id)
    for r in requests:
        rid = r["request_id"]
        linked = [m for m in all_markups if m.get("request_id") == rid]
        r["markup_count"] = len(linked)
        r["open_markup_count"] = sum(1 for m in linked if m.get("status") == "open")
        r["confirmed_markup_count"] = sum(1 for m in linked if m.get("status") == "confirmed")

        # Backward compat: ensure new fields exist
        r.setdefault("lead_reviewer", r.get("to_name", ""))
        r.setdefault("squad_reviewers", [])
        r.setdefault("due_date", "")
        r.setdefault("return_code", "")
        r.setdefault("transmittal_no", "")
        r.setdefault("reviewer_statuses", {})

    return {"status": "success", "requests": requests}


@router.patch("/projects/{project_id}/requests/{request_id}")
async def update_review_request_status(
    project_id: str,
    request_id: str,
    req: UpdateReviewRequestStatusModel,
    authorization: Optional[str] = Header(None)
):
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    review_req = fs_get_request(project_id, request_id)
    if not review_req:
        raise HTTPException(status_code=404, detail="Request not found")

    now = datetime.now(timezone.utc).isoformat()
    fs_update_request(project_id, request_id, {
        "status": req.status,
        "updated_at": now,
    })

    if req.status == "confirmed":
        fs_add_activity(project_id, "request_confirmed", {
            "request_id": request_id,
        }, actor=username)

    review_req["status"] = req.status
    review_req["updated_at"] = now
    return {"status": "success", "request": review_req}


@router.post("/projects/{project_id}/requests/{request_id}/replies")
async def add_review_request_reply(
    project_id: str,
    request_id: str,
    req: ReplyReviewRequestModel,
    authorization: Optional[str] = Header(None)
):
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    review_req = fs_get_request(project_id, request_id)
    if not review_req:
        raise HTTPException(status_code=404, detail="Request not found")

    now = datetime.now(timezone.utc).isoformat()
    reply = {
        "reply_id": str(uuid.uuid4())[:8],
        "author_name": req.author_name or username,
        "content": req.content,
        "created_at": now,
    }

    replies = review_req.get("replies", [])
    replies.append(reply)

    fs_update_request(project_id, request_id, {
        "replies": replies,
        "updated_at": now,
    })

    return {"status": "success", "reply": reply}


@router.delete("/projects/{project_id}/requests/{request_id}")
async def delete_review_request(
    project_id: str,
    request_id: str,
    authorization: Optional[str] = Header(None)
):
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    fs_delete_request(project_id, request_id)
    return {"status": "success", "deleted": request_id}


# ── EPC Workflow: Intake Decision ──

@router.post("/projects/{project_id}/requests/{request_id}/intake-decision")
async def intake_decision(
    project_id: str,
    request_id: str,
    req: IntakeDecisionRequest,
    authorization: Optional[str] = Header(None)
):
    """Accept or reject a drawing at intake stage."""
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    now = datetime.now(timezone.utc).isoformat()

    # Update drawing intake_status
    drawing = fs_get_drawing(project_id, req.drawing_id)
    if drawing:
        fs_update_drawing(project_id, req.drawing_id, {
            "intake_status": req.decision,
            "vdrl_match": req.vdrl_match if req.vdrl_match is not None else False,
            "intake_comment": req.comment or "",
            "updated_at": now,
        })
        fs_update_project(project_id, {"updated_at": now})

    # Update request status
    review_req = fs_get_request(project_id, request_id)
    if not review_req:
        raise HTTPException(status_code=404, detail="Request not found")

    new_status = "intake" if req.decision == "accepted" else "rejected"
    fs_update_request(project_id, request_id, {
        "status": new_status,
        "updated_at": now,
    })

    fs_add_activity(project_id, "intake_decision", {
        "request_id": request_id,
        "drawing_id": req.drawing_id,
        "decision": req.decision,
    }, actor=username)

    review_req["status"] = new_status
    review_req["updated_at"] = now
    return {"status": "success", "request": review_req}


# ── EPC Workflow: Assign Reviewers ──

@router.post("/projects/{project_id}/requests/{request_id}/assign")
async def assign_reviewers(
    project_id: str,
    request_id: str,
    req: AssignReviewersRequest,
    authorization: Optional[str] = Header(None)
):
    """Assign lead + squad reviewers and set due date."""
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    review_req = fs_get_request(project_id, request_id)
    if not review_req:
        raise HTTPException(status_code=404, detail="Request not found")

    now = datetime.now(timezone.utc).isoformat()

    # Build reviewer_statuses
    reviewer_statuses = {
        req.lead_reviewer: {"role": "lead", "status": "pending", "completed_at": None}
    }
    for sr in (req.squad_reviewers or []):
        reviewer_statuses[sr] = {"role": "squad", "status": "pending", "completed_at": None}

    # Due date: use provided or default to 14 days from now
    due_date = req.due_date
    if not due_date and not review_req.get("due_date"):
        from datetime import timedelta
        due = datetime.now(timezone.utc) + timedelta(days=14)
        due_date = due.strftime("%Y-%m-%d")
    elif not due_date:
        due_date = review_req.get("due_date", "")

    updates = {
        "lead_reviewer": req.lead_reviewer,
        "squad_reviewers": req.squad_reviewers or [],
        "to_name": req.lead_reviewer,  # backward compat
        "reviewer_statuses": reviewer_statuses,
        "due_date": due_date,
        "status": "assigned",
        "updated_at": now,
    }

    fs_update_request(project_id, request_id, updates)

    fs_add_activity(project_id, "reviewers_assigned", {
        "request_id": request_id,
        "lead_reviewer": req.lead_reviewer,
        "squad_reviewers": req.squad_reviewers or [],
    }, actor=username)

    review_req.update(updates)
    return {"status": "success", "request": review_req}


# ── EPC Workflow: Update Reviewer Status ──

@router.patch("/projects/{project_id}/requests/{request_id}/reviewer-status")
async def update_reviewer_status(
    project_id: str,
    request_id: str,
    reviewer_name: str = Query(...),
    new_status: str = Query(...),  # "in_progress" | "done"
    authorization: Optional[str] = Header(None)
):
    """Update individual reviewer's status within a request."""
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    review_req = fs_get_request(project_id, request_id)
    if not review_req:
        raise HTTPException(status_code=404, detail="Request not found")

    now = datetime.now(timezone.utc).isoformat()
    reviewer_statuses = review_req.get("reviewer_statuses", {})

    if reviewer_name in reviewer_statuses:
        reviewer_statuses[reviewer_name]["status"] = new_status
        if new_status == "done":
            reviewer_statuses[reviewer_name]["completed_at"] = now
    else:
        reviewer_statuses[reviewer_name] = {"role": "squad", "status": new_status, "completed_at": now if new_status == "done" else None}

    # Check if all reviewers are done -> auto-update request status
    all_done = all(rs.get("status") == "done" for rs in reviewer_statuses.values())
    req_status = "markup_done" if (all_done and reviewer_statuses) else review_req.get("status")

    fs_update_request(project_id, request_id, {
        "reviewer_statuses": reviewer_statuses,
        "status": req_status,
        "updated_at": now,
    })

    review_req["reviewer_statuses"] = reviewer_statuses
    review_req["status"] = req_status
    review_req["updated_at"] = now
    return {"status": "success", "request": review_req}


# ── EPC Workflow: Consolidate ──

@router.post("/projects/{project_id}/requests/{request_id}/consolidate")
async def consolidate_review(
    project_id: str,
    request_id: str,
    req: ConsolidateRequest,
    authorization: Optional[str] = Header(None)
):
    """Lead consolidates all markups. Moves to consolidation status."""
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    review_req = fs_get_request(project_id, request_id)
    if not review_req:
        raise HTTPException(status_code=404, detail="Request not found")

    # Verify all reviewers are done
    reviewer_statuses = review_req.get("reviewer_statuses", {})
    not_done = [name for name, rs in reviewer_statuses.items() if rs.get("status") != "done"]
    if not_done:
        raise HTTPException(status_code=400, detail=f"Reviewers not done: {', '.join(not_done)}")

    now = datetime.now(timezone.utc).isoformat()

    # Mark selected markups as final
    if req.confirmed_markup_ids:
        fs_bulk_update_markups(project_id, req.confirmed_markup_ids, {
            "status": "final",
            "updated_at": now,
        })

    fs_update_request(project_id, request_id, {
        "status": "consolidation",
        "updated_at": now,
    })

    fs_add_activity(project_id, "review_consolidated", {
        "request_id": request_id,
        "confirmed_markups": len(req.confirmed_markup_ids or []),
    }, actor=username)

    review_req["status"] = "consolidation"
    review_req["updated_at"] = now
    return {"status": "success", "request": review_req}


# ── EPC Workflow: Conflicts Detection ──

@router.get("/projects/{project_id}/requests/{request_id}/conflicts")
async def get_conflicts(
    project_id: str,
    request_id: str,
    authorization: Optional[str] = Header(None)
):
    """Detect markup conflicts (different disciplines marking same area)."""
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    review_req = fs_get_request(project_id, request_id)
    if not review_req:
        raise HTTPException(status_code=404, detail="Request not found")

    # Filter markups linked to this request's drawing
    drawing_id = review_req.get("drawing_id", "")
    drawing_markups = fs_list_markups(project_id, drawing_id=drawing_id)

    # Detect conflicts: markups on same page, different disciplines, within proximity
    conflicts = []
    proximity = 0.05  # 5% of page dimension
    for i, m1 in enumerate(drawing_markups):
        for m2 in drawing_markups[i+1:]:
            if m1.get("page") != m2.get("page"):
                continue
            if m1.get("discipline") == m2.get("discipline"):
                continue
            dx = abs(m1.get("x", 0) - m2.get("x", 0))
            dy = abs(m1.get("y", 0) - m2.get("y", 0))
            dist = math.sqrt(dx*dx + dy*dy)
            if dist <= proximity:
                conflicts.append({
                    "markup_a": {"markup_id": m1["markup_id"], "discipline": m1.get("discipline"), "comment": m1.get("comment", ""), "author_name": m1.get("author_name", "")},
                    "markup_b": {"markup_id": m2["markup_id"], "discipline": m2.get("discipline"), "comment": m2.get("comment", ""), "author_name": m2.get("author_name", "")},
                    "page": m1.get("page"),
                    "distance": round(dist, 4),
                })

    return {"status": "success", "conflicts": conflicts, "count": len(conflicts)}


# ── EPC Workflow: Return Code ──

@router.post("/projects/{project_id}/requests/{request_id}/return-code")
async def set_return_code(
    project_id: str,
    request_id: str,
    req: ReturnCodeRequest,
    authorization: Optional[str] = Header(None)
):
    """Set return code and update related statuses."""
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    valid_codes = ["code_1", "code_2", "code_3", "code_4"]
    if req.return_code not in valid_codes:
        raise HTTPException(status_code=400, detail=f"Invalid return code. Must be one of: {valid_codes}")

    review_req = fs_get_request(project_id, request_id)
    if not review_req:
        raise HTTPException(status_code=404, detail="Request not found")

    now = datetime.now(timezone.utc).isoformat()
    fs_update_request(project_id, request_id, {
        "return_code": req.return_code,
        "status": "return_decided",
        "updated_at": now,
    })

    # Update drawing em_approval based on return code
    drawing_id = review_req.get("drawing_id")
    if drawing_id:
        drawing = fs_get_drawing(project_id, drawing_id)
        if drawing:
            code_to_approval = {
                "code_1": "approved",
                "code_2": "conditionally_approved",
                "code_3": "rejected",
                "code_4": "info_only",
            }
            em_approval = {
                "status": code_to_approval.get(req.return_code, "pending"),
                "approver_name": username,
                "comment": req.comment or "",
                "approved_at": now,
            }

            drawing_updates = {
                "em_approval": em_approval,
                "updated_at": now,
            }

            if req.return_code == "code_3":
                drawing_updates["staging_status"] = "resubmit_required"

            # Update review_status for all disciplines
            code_to_review = {
                "code_1": "completed",
                "code_2": "completed",
                "code_3": "rejected",
                "code_4": "completed",
            }
            review_status = drawing.get("review_status", {})
            for disc in review_status:
                review_status[disc]["status"] = code_to_review.get(req.return_code, review_status[disc].get("status"))
                review_status[disc]["updated_at"] = now
            drawing_updates["review_status"] = review_status

            fs_update_drawing(project_id, drawing_id, drawing_updates)
            fs_update_project(project_id, {"updated_at": now})

    fs_add_activity(project_id, "return_code_set", {
        "request_id": request_id,
        "return_code": req.return_code,
        "drawing_id": review_req.get("drawing_id", ""),
    }, actor=username)

    review_req["return_code"] = req.return_code
    review_req["status"] = "return_decided"
    review_req["updated_at"] = now
    return {"status": "success", "request": review_req}


# ── EPC Workflow: Transmittal ──

@router.post("/projects/{project_id}/requests/{request_id}/transmittal")
async def create_transmittal(
    project_id: str,
    request_id: str,
    req: TransmittalRequest,
    authorization: Optional[str] = Header(None)
):
    """Generate transmittal number, update status to transmitted."""
    username = _get_username(authorization)
    project = fs_resolve_project(username, project_id)

    review_req = fs_get_request(project_id, request_id)
    if not review_req:
        raise HTTPException(status_code=404, detail="Request not found")

    now = datetime.now(timezone.utc).isoformat()

    # Generate transmittal number
    project_code = project.get("project_code", "PROJ") or "PROJ"
    existing_transmittals = fs_count_transmittals(project_id)
    tr_number = f"{project_code}-TR-{existing_transmittals + 1:03d}"

    fs_update_request(project_id, request_id, {
        "transmittal_no": tr_number,
        "status": "transmitted",
        "transmitted_at": now,
        "updated_at": now,
    })

    # Update drawing revision status
    drawing_id = review_req.get("drawing_id")
    if drawing_id:
        drawing = fs_get_drawing(project_id, drawing_id)
        if drawing:
            fs_update_drawing(project_id, drawing_id, {"updated_at": now})
            fs_update_project(project_id, {"updated_at": now})

    fs_add_activity(project_id, "transmittal_created", {
        "request_id": request_id,
        "transmittal_no": tr_number,
        "drawing_id": review_req.get("drawing_id", ""),
        "return_code": review_req.get("return_code", ""),
    }, actor=username)

    review_req["transmittal_no"] = tr_number
    review_req["status"] = "transmitted"
    review_req["transmitted_at"] = now
    review_req["updated_at"] = now

    print(f"[PlantSync] Transmittal created: {tr_number} for request {request_id}", flush=True)
    return {"status": "success", "request": review_req, "transmittal_no": tr_number}


# ── Feature 2: Staging Area ──

class RegisterDrawingRequest(BaseModel):
    drawing_number: Optional[str] = None
    title: Optional[str] = None
    revision: Optional[str] = None
    discipline: Optional[str] = None
    vendor_drawing_number: Optional[str] = None
    issue_purpose: Optional[str] = None
    issue_date: Optional[str] = None
    receive_date: Optional[str] = None
    vendor_name: Optional[str] = None
    reviewer_name: Optional[str] = None
    has_dwg: Optional[bool] = None
    related_drawings: Optional[List[str]] = None
    change_log: Optional[str] = None
    remarks: Optional[str] = None


@router.post("/projects/{project_id}/drawings/{drawing_id}/register")
async def register_drawing(
    project_id: str,
    drawing_id: str,
    req: RegisterDrawingRequest,
    authorization: Optional[str] = Header(None)
):
    """Move drawing from staged → registered, update metadata."""
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    drawing = fs_get_drawing(project_id, drawing_id)
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing not found")

    updates = {"staging_status": "registered"}

    # Update all provided fields
    for field in [
        "drawing_number", "title", "discipline",
        "vendor_drawing_number", "issue_purpose", "issue_date", "receive_date",
        "vendor_name", "reviewer_name", "has_dwg", "change_log", "remarks",
    ]:
        val = getattr(req, field, None)
        if val is not None:
            updates[field] = val

    if req.revision is not None:
        updates["current_revision"] = req.revision
        revisions = drawing.get("revisions", [])
        if revisions:
            revisions[-1]["revision"] = req.revision
            updates["revisions"] = revisions

    if req.related_drawings is not None:
        updates["related_drawings"] = req.related_drawings

    now = datetime.now(timezone.utc).isoformat()
    updates["updated_at"] = now
    fs_update_drawing(project_id, drawing_id, updates)
    fs_update_project(project_id, {"updated_at": now})

    fs_add_activity(project_id, "drawing_registered", {
        "drawing_id": drawing_id,
        "drawing_number": updates.get("drawing_number", drawing.get("drawing_number", "")),
    }, actor=username)

    drawing.update(updates)
    print(f"[PlantSync] Drawing registered: {drawing_id}", flush=True)
    return {"status": "success", "drawing": drawing}


@router.get("/projects/{project_id}/staged")
async def list_staged_drawings(
    project_id: str,
    authorization: Optional[str] = Header(None)
):
    """Return only staged (not yet registered) drawings."""
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    drawings = fs_list_drawings(project_id)
    staged = [d for d in drawings if d.get("staging_status") == "staged"]
    return {"status": "success", "staged": staged, "count": len(staged)}


# ── Feature 3: Visual Diff Viewer ──

class DiffUrlsRequest(BaseModel):
    revision_id_a: str
    revision_id_b: str


@router.post("/projects/{project_id}/drawings/{drawing_id}/diff-urls")
async def get_diff_urls(
    project_id: str,
    drawing_id: str,
    req: DiffUrlsRequest,
    authorization: Optional[str] = Header(None)
):
    """Return SAS URLs for two revisions to enable visual diff."""
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    drawing = fs_get_drawing(project_id, drawing_id)
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing not found")

    revisions = drawing.get("revisions", [])
    rev_a = next((r for r in revisions if r["revision_id"] == req.revision_id_a), None)
    rev_b = next((r for r in revisions if r["revision_id"] == req.revision_id_b), None)

    if not rev_a or not rev_b:
        raise HTTPException(status_code=404, detail="One or both revisions not found")

    from app.services.blob_storage import generate_sas_url
    url_a = generate_sas_url(rev_a["blob_path"])
    url_b = generate_sas_url(rev_b["blob_path"])

    return {
        "status": "success",
        "revision_a": {**rev_a, "pdf_url": url_a},
        "revision_b": {**rev_b, "pdf_url": url_b},
    }


# ── Feature 4: Smart Markup Pin + AI 제안 ──

class NearbyTextRequest(BaseModel):
    page: int
    x: float
    y: float
    radius: float = 0.05


@router.post("/projects/{project_id}/drawings/{drawing_id}/nearby-text")
async def get_nearby_text(
    project_id: str,
    drawing_id: str,
    req: NearbyTextRequest,
    authorization: Optional[str] = Header(None)
):
    """Find DI words/lines near a pin coordinate on a specific page."""
    username = _get_username(authorization)
    project = fs_resolve_project(username, project_id)
    owner = project["owner"]
    container = _get_container()

    drawing = fs_get_drawing(project_id, drawing_id)
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing not found")

    # Find di_result.json (still on blob)
    revisions = drawing.get("revisions", [])
    if not revisions:
        return {"status": "success", "words": [], "lines": []}

    latest_rev = revisions[-1]
    blob_dir = latest_rev["blob_path"].rsplit("/", 1)[0]
    di_path = f"{blob_dir}/di_result.json"
    di_result = _load_json(container, di_path)
    if not di_result:
        return {"status": "success", "words": [], "lines": []}

    # Get the requested page (0-indexed in di_result)
    page_idx = req.page - 1
    if page_idx < 0 or page_idx >= len(di_result):
        return {"status": "success", "words": [], "lines": []}

    page_data = di_result[page_idx]
    page_w = page_data.get("width", 1)
    page_h = page_data.get("height", 1)

    # Pin coords: normalized (0-1) → DI pixel space
    pin_x = req.x * page_w
    pin_y = req.y * page_h
    radius_px = req.radius * max(page_w, page_h)

    def polygon_center(polygon):
        if not polygon or len(polygon) < 4:
            return None, None
        xs = [polygon[i] for i in range(0, len(polygon), 2)]
        ys = [polygon[i] for i in range(1, len(polygon), 2)]
        return sum(xs) / len(xs), sum(ys) / len(ys)

    def dist(cx, cy):
        return math.sqrt((cx - pin_x) ** 2 + (cy - pin_y) ** 2)

    # Words
    nearby_words = []
    for w in page_data.get("words", []):
        cx, cy = polygon_center(w.get("polygon", []))
        if cx is None:
            continue
        d = dist(cx, cy)
        if d <= radius_px:
            nearby_words.append({
                "content": w.get("content", ""),
                "confidence": w.get("confidence", 0),
                "distance": round(d, 2),
            })
    nearby_words.sort(key=lambda x: x["distance"])

    # Lines
    nearby_lines = []
    for line in page_data.get("lines", []):
        cx, cy = polygon_center(line.get("polygon", []))
        if cx is None:
            continue
        d = dist(cx, cy)
        if d <= radius_px * 2:  # larger radius for lines
            nearby_lines.append({
                "content": line.get("content", ""),
                "distance": round(d, 2),
            })
    nearby_lines.sort(key=lambda x: x["distance"])

    return {
        "status": "success",
        "words": nearby_words[:20],
        "lines": nearby_lines[:10],
    }


class RelatedSearchRequest(BaseModel):
    query: str


@router.post("/projects/{project_id}/drawings/{drawing_id}/related-search")
async def related_search(
    project_id: str,
    drawing_id: str,
    req: RelatedSearchRequest,
    authorization: Optional[str] = Header(None)
):
    """Search related markups in the project and Azure AI Search index."""
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    query = req.query.strip().lower()
    if not query:
        return {"status": "success", "markups": [], "documents": []}

    # 1) Search markups for keyword matches
    all_markups = fs_list_markups_all(project_id)
    matched_markups = []
    for m in all_markups:
        text = (m.get("comment", "") + " " + " ".join(
            r.get("content", "") for r in m.get("replies", [])
        )).lower()
        if query in text:
            matched_markups.append({
                "markup_id": m.get("markup_id"),
                "drawing_id": m.get("drawing_id"),
                "comment": m.get("comment", ""),
                "discipline": m.get("discipline", ""),
                "author_name": m.get("author_name", ""),
                "status": m.get("status", ""),
                "page": m.get("page", 1),
            })

    # 2) Search Azure AI Search (best-effort, skip if not configured)
    search_results = []
    try:
        from app.services.azure_search import search_documents
        hits = search_documents(query, top=5)
        for h in hits:
            search_results.append({
                "type": "document",
                "title": h.get("title", ""),
                "content_snippet": h.get("content", "")[:200],
                "score": h.get("@search.score", 0),
                "source": h.get("source", ""),
            })
    except Exception as e:
        logger.debug(f"Azure Search skipped: {e}")

    return {
        "status": "success",
        "markups": matched_markups[:10],
        "documents": search_results,
    }


# ── Feature 5: Excel Export ──

@router.get("/projects/{project_id}/export-excel")
async def export_excel(
    project_id: str,
    authorization: Optional[str] = Header(None)
):
    """Export drawing register and markup details as Excel."""
    username = _get_username(authorization)
    project = fs_resolve_project(username, project_id)

    drawings = fs_list_drawings(project_id)
    markups = fs_list_markups_all(project_id)

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl is not installed")

    wb = openpyxl.Workbook()

    # ── Sheet 1: 도면 대장 ──
    ws1 = wb.active
    ws1.title = "도면 대장"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers1 = [
        ("No", 5), ("도면번호", 25), ("타이틀", 40), ("리비전", 8),
        ("디시플린", 10), ("Issue Purpose", 15), ("Vendor", 15),
        ("등록일", 15),
        ("공정", 8), ("기계", 8), ("배관", 8), ("전기", 8), ("계장", 8), ("토목", 8),
        ("EM 승인", 10), ("마크업수", 8),
    ]
    for col_idx, (header, width) in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    cell_align = Alignment(vertical="top", wrap_text=True)
    for row_idx, d in enumerate(drawings, 2):
        rs = d.get("review_status", {})
        markup_count = sum(1 for m in markups if m.get("drawing_id") == d.get("drawing_id"))
        values = [
            row_idx - 1,
            d.get("drawing_number", ""),
            d.get("title", ""),
            d.get("current_revision", ""),
            d.get("discipline", ""),
            d.get("issue_purpose", ""),
            d.get("vendor_name", ""),
            d.get("created_at", "")[:10] if d.get("created_at") else "",
            rs.get("process", {}).get("status", ""),
            rs.get("mechanical", {}).get("status", ""),
            rs.get("piping", {}).get("status", ""),
            rs.get("electrical", {}).get("status", ""),
            rs.get("instrument", {}).get("status", ""),
            rs.get("civil", {}).get("status", ""),
            d.get("em_approval", {}).get("status", "pending"),
            markup_count,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = cell_align
            cell.border = thin_border

    # ── Sheet 2: 마크업 내역 ──
    ws2 = wb.create_sheet("마크업 내역")
    headers2 = [
        ("No", 5), ("도면번호", 25), ("페이지", 6), ("디시플린", 10),
        ("코멘트", 50), ("작성자", 12), ("상태", 8), ("생성일", 18),
    ]
    for col_idx, (header, width) in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Build drawing_id → drawing_number map
    did_to_num = {d["drawing_id"]: d.get("drawing_number", "") for d in drawings}

    for row_idx, m in enumerate(markups, 2):
        values = [
            row_idx - 1,
            did_to_num.get(m.get("drawing_id", ""), ""),
            m.get("page", ""),
            m.get("discipline", ""),
            m.get("comment", ""),
            m.get("author_name", ""),
            m.get("status", ""),
            m.get("created_at", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = cell_align
            cell.border = thin_border

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    project_name = project.get("project_name", "plantsync")
    safe_filename = f"{project_name}_도면대장.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_filename}"'},
    )


# ── Feature 6: Bulk Upload ──

@router.post("/projects/{project_id}/bulk-upload")
async def bulk_upload_drawings(
    project_id: str,
    files: List[UploadFile] = File(...),
    authorization: Optional[str] = Header(None)
):
    """Upload multiple drawing PDFs at once. Each file goes through the same DI → title block → staging flow."""
    username = _get_username(authorization)
    project = fs_resolve_project(username, project_id)
    owner = project["owner"]
    container = _get_container()

    results = []
    for file in files:
        try:
            file_content = await file.read()
            if not file_content:
                results.append({"filename": file.filename, "status": "error", "error": "Empty file"})
                continue

            drawing_id = str(uuid.uuid4())[:8]
            filename = file.filename or f"{drawing_id}.pdf"

            # Upload PDF to temp path
            temp_blob_path = f"{owner}/plantsync/{project_id}/temp/{drawing_id}/{filename}"
            container.upload_blob(name=temp_blob_path, data=file_content, overwrite=True)

            # Generate SAS URL for DI analysis
            from app.services.blob_storage import generate_sas_url
            pdf_url = generate_sas_url(temp_blob_path)

            # Run Document Intelligence
            di_result = []
            title_block = {"drawing_number": "", "title": "", "revision": "", "discipline": "",
                            "issue_purpose": "", "issue_date": "", "vendor_drawing_number": "", "vendor_name": ""}
            try:
                from app.services.azure_di import azure_di_service
                di_result = azure_di_service.analyze_document_from_url(pdf_url)
                title_block = _extract_title_block(di_result)
            except Exception as e:
                print(f"[PlantSync] DI analysis failed for {filename}: {e}", flush=True)

            discipline = title_block.get("discipline") or "unknown"

            # Move PDF to discipline folder
            blob_path = f"{owner}/plantsync/{project_id}/{discipline}/{drawing_id}/{filename}"
            container.upload_blob(name=blob_path, data=file_content, overwrite=True)
            try:
                container.delete_blob(temp_blob_path)
            except Exception:
                pass

            # Save DI result (still as blob)
            if di_result:
                di_path = f"{owner}/plantsync/{project_id}/{discipline}/{drawing_id}/di_result.json"
                _save_json(container, di_path, di_result)

            page_count = len(di_result) if di_result else 1
            now = datetime.now(timezone.utc).isoformat()

            # Check for existing drawing with same number
            existing_drawing = None
            if title_block.get("drawing_number"):
                existing_drawing = fs_find_drawing_by_number(project_id, title_block["drawing_number"])

            if existing_drawing:
                rev_entry = {
                    "revision_id": drawing_id,
                    "revision": title_block.get("revision", ""),
                    "blob_path": blob_path,
                    "uploaded_at": now,
                    "page_count": page_count,
                    "filename": filename,
                }
                existing_drawing["revisions"].append(rev_entry)
                fs_update_drawing(project_id, existing_drawing["drawing_id"], {
                    "revisions": existing_drawing["revisions"],
                    "current_revision": title_block.get("revision", existing_drawing.get("current_revision", "")),
                    "updated_at": now,
                })
            else:
                today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
                drawing_data = {
                    "drawing_id": drawing_id,
                    "drawing_number": title_block.get("drawing_number", ""),
                    "title": title_block.get("title", filename),
                    "discipline": discipline,
                    "current_revision": title_block.get("revision", "A"),
                    "revisions": [{
                        "revision_id": drawing_id,
                        "revision": title_block.get("revision", "A"),
                        "blob_path": blob_path,
                        "uploaded_at": now,
                        "page_count": page_count,
                        "filename": filename,
                    }],
                    "review_status": _empty_review_status(),
                    "em_approval": {"status": "pending"},
                    "staging_status": "staged",
                    "vendor_drawing_number": title_block.get("vendor_drawing_number", ""),
                    "issue_purpose": title_block.get("issue_purpose", ""),
                    "issue_date": title_block.get("issue_date", ""),
                    "receive_date": today_str,
                    "vendor_name": title_block.get("vendor_name", ""),
                    "reviewer_name": "",
                    "has_dwg": False,
                    "related_drawings": [],
                    "change_log": "",
                    "remarks": "",
                    "created_at": now,
                    "updated_at": now,
                }
                fs_add_drawing(project_id, drawing_id, drawing_data)

            results.append({
                "filename": filename,
                "status": "success",
                "drawing_id": drawing_id,
                "drawing_number": title_block.get("drawing_number", ""),
                "title": title_block.get("title", ""),
                "discipline": discipline,
                "is_new_revision": existing_drawing is not None,
            })

        except Exception as e:
            results.append({"filename": file.filename, "status": "error", "error": str(e)})

    fs_update_project(project_id, {"updated_at": datetime.now(timezone.utc).isoformat()})

    success_count = sum(1 for r in results if r["status"] == "success")
    fs_add_activity(project_id, "drawing_uploaded", {
        "bulk": True,
        "total": len(files),
        "success": success_count,
    }, actor=username)

    print(f"[PlantSync] Bulk upload: {success_count}/{len(files)} succeeded", flush=True)
    return {"status": "success", "results": results, "success_count": success_count, "total": len(files)}


# ── Feature 7: Activity Timeline ──

@router.get("/projects/{project_id}/activity")
async def get_activity(
    project_id: str,
    limit: int = Query(50, ge=1, le=200),
    authorization: Optional[str] = Header(None)
):
    """Return recent activity events for the project."""
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    activities = fs_list_activities(project_id, limit=limit)
    return {"status": "success", "activities": activities}


# ── Feature 8: Review Gate ──

@router.get("/projects/{project_id}/drawings/{drawing_id}/review-gate")
async def review_gate(
    project_id: str,
    drawing_id: str,
    authorization: Optional[str] = Header(None)
):
    """Check if all discipline reviews are completed for EM approval."""
    username = _get_username(authorization)
    fs_resolve_project(username, project_id)

    drawing = fs_get_drawing(project_id, drawing_id)
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing not found")

    review_status = drawing.get("review_status", _empty_review_status())

    completed = []
    incomplete = []
    for disc in DISCIPLINES:
        rs = review_status.get(disc, {})
        st = rs.get("status", "not_started")
        if st == "completed":
            completed.append(disc)
        else:
            incomplete.append({"discipline": disc, "status": st})

    total = len(DISCIPLINES)
    completion_rate = round(len(completed) / total * 100) if total > 0 else 0
    all_completed = len(incomplete) == 0

    return {
        "status": "success",
        "all_completed": all_completed,
        "completion_rate": completion_rate,
        "completed_disciplines": completed,
        "incomplete_disciplines": incomplete,
        "total": total,
    }


# ── Feature 9: Markup PDF Export ──

@router.post("/projects/{project_id}/drawings/{drawing_id}/export-markup-pdf")
async def export_markup_pdf(
    project_id: str,
    drawing_id: str,
    authorization: Optional[str] = Header(None)
):
    """Render markup pins on the original PDF and return as download."""
    username = _get_username(authorization)
    project = fs_resolve_project(username, project_id)
    container = _get_container()

    drawing = fs_get_drawing(project_id, drawing_id)
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing not found")

    # Get latest revision PDF
    revisions = drawing.get("revisions", [])
    if not revisions:
        raise HTTPException(status_code=404, detail="No revisions found")

    latest_rev = revisions[-1]
    blob_path = latest_rev["blob_path"]

    # Download PDF from blob
    try:
        blob_client = container.get_blob_client(blob_path)
        pdf_bytes = blob_client.download_blob().readall()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to download PDF: {e}")

    # Load markups for this drawing
    drawing_markups = fs_list_markups(project_id, drawing_id=drawing_id)

    if not drawing_markups:
        # No markups — return original PDF
        buffer = io.BytesIO(pdf_bytes)
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{drawing.get("drawing_number", drawing_id)}_markup.pdf"'},
        )

    try:
        import fitz  # PyMuPDF
    except ImportError:
        raise HTTPException(status_code=500, detail="PyMuPDF (fitz) is not installed")

    DISC_COLORS = {
        "process":     (0.94, 0.27, 0.27),  # red
        "mechanical":  (0.23, 0.51, 0.96),  # blue
        "piping":      (0.13, 0.77, 0.37),  # green
        "electrical":  (0.92, 0.70, 0.03),  # yellow
        "instrument":  (0.66, 0.33, 0.97),  # purple
        "civil":       (0.98, 0.45, 0.09),  # orange
    }

    doc = fitz.open(stream=pdf_bytes, filetype="pdf")

    # Group markups by page
    markups_by_page = {}
    for m in drawing_markups:
        pg = m.get("page", 1)
        markups_by_page.setdefault(pg, []).append(m)

    for page_num, page_markups in markups_by_page.items():
        page_idx = page_num - 1
        if page_idx < 0 or page_idx >= len(doc):
            continue

        page = doc[page_idx]
        rect = page.rect
        pw, ph = rect.width, rect.height

        for idx, m in enumerate(page_markups, 1):
            x = m.get("x", 0) * pw
            y = m.get("y", 0) * ph
            disc = m.get("discipline", "")
            color = DISC_COLORS.get(disc, (0.5, 0.5, 0.5))

            # Draw filled circle
            center = fitz.Point(x, y)
            r = 12
            circle_rect = fitz.Rect(x - r, y - r, x + r, y + r)
            shape = page.new_shape()
            shape.draw_circle(center, r)
            shape.finish(color=color, fill=color, fill_opacity=0.85)
            shape.commit()

            # Draw number text inside circle
            fontsize = 9
            text = str(idx)
            text_point = fitz.Point(x - fontsize * 0.3 * len(text), y + fontsize * 0.35)
            page.insert_text(text_point, text, fontsize=fontsize, color=(1, 1, 1))

            # Add annotation comment as a text note
            comment = m.get("comment", "")
            author = m.get("author_name", "")
            status = m.get("status", "")
            note_text = f"[{disc}] {comment}\n- {author} ({status})"

            annot = page.add_text_annot(center, note_text)
            annot.set_colors(stroke=color)
            annot.update()

    # Save to buffer
    output = io.BytesIO()
    doc.save(output)
    doc.close()
    output.seek(0)

    safe_name = drawing.get("drawing_number", drawing_id)

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_markup.pdf"'},
    )
