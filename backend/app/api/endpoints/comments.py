"""
PDF Comment Extractor - API Endpoints

- POST /extract       : Upload PDF → extract annotations → return JSON
- POST /extract-blob  : Download PDF from Azure Blob → extract annotations → return JSON
- POST /delete-blob   : Delete comment PDF + JSON + search index
- POST /export-excel  : Receive table data → return Excel file
"""

import io
import os
import re
import logging
from datetime import datetime
from typing import List, Optional

import fitz  # PyMuPDF
from fastapi import APIRouter, HTTPException, UploadFile, File, Body
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

logger = logging.getLogger(__name__)
router = APIRouter()


# ── Models ──

class CommentRow(BaseModel):
    no: int
    drawing_no: str = ""
    page: int = 0
    type: str = ""
    author: str = ""
    contents: str = ""
    reply: str = ""
    created_date: str = ""


class ExportRequest(BaseModel):
    filename: str = "comments.xlsx"
    rows: List[CommentRow]


# ── Helpers ──

def _extract_drawing_no(filename: str) -> str:
    """Try to extract a drawing number from the PDF filename."""
    name = filename.rsplit(".", 1)[0] if "." in filename else filename
    # Common patterns: XXX-XXXX-XXX or similar alphanumeric-dash codes
    match = re.search(r'[\dA-Z][\dA-Z\-]{4,}[\dA-Z]', name, re.IGNORECASE)
    return match.group(0) if match else name


def _parse_date(date_str: Optional[str]) -> str:
    """Parse PDF date string (D:YYYYMMDDHHmmSS) to readable format."""
    if not date_str:
        return ""
    try:
        # Strip 'D:' prefix and timezone info
        clean = date_str.replace("D:", "").split("+")[0].split("-")[0].split("Z")[0]
        if len(clean) >= 8:
            dt = datetime.strptime(clean[:14].ljust(14, "0"), "%Y%m%d%H%M%S")
            return dt.strftime("%Y-%m-%d %H:%M")
        return clean
    except Exception:
        return date_str or ""


def _split_comments(text: str) -> list[str]:
    """Split multi-comment text into individual comments by code pattern (e.g., G1., M2., E1.)."""
    if not text or not text.strip():
        return [text or ""]
    # Lookahead split on comment codes: 1-3 uppercase letters + 1-3 digits + dot
    # e.g., G1., M2., M10., E1., P1.
    parts = re.split(r'(?=[A-Z]{1,3}\d{1,3}\.\s)', text)
    results = []
    for part in parts:
        cleaned = part.replace('\r\n', ' ').replace('\r', ' ').replace('\n', ' ').strip()
        if cleaned:
            results.append(cleaned)
    return results if results else [text.strip()]


ANNOT_TYPE_MAP = {
    0: "Text",
    1: "Link",
    2: "FreeText",
    3: "Line",
    4: "Square",
    5: "Circle",
    6: "Polygon",
    7: "PolyLine",
    8: "Highlight",
    9: "Underline",
    10: "Squiggly",
    11: "StrikeOut",
    12: "Stamp",
    13: "Caret",
    14: "Ink",
    15: "Popup",
    16: "FileAttachment",
    17: "Sound",
    18: "Movie",
    19: "Widget",
    20: "Screen",
    21: "PrinterMark",
    22: "TrapNet",
    23: "Watermark",
    24: "3D",
    25: "Redact",
}


def _extract_annotations(pdf_bytes: bytes, filename: str) -> dict:
    """PyMuPDF로 PDF 어노테이션 추출 (공통 로직)"""
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    except Exception as e:
        logger.error(f"Failed to open PDF: {e}", exc_info=True)
        raise HTTPException(status_code=400, detail=f"PDF 파일을 열 수 없습니다: {str(e)}")

    drawing_no = _extract_drawing_no(filename)
    total_pages = len(doc)
    comments = []
    idx = 1

    for page_num in range(total_pages):
        page = doc[page_num]
        for annot in page.annots() or []:
            annot_type = annot.type[0] if annot.type else -1
            # Skip Link and Popup annotations as they're not user comments
            if annot_type in (1, 15):
                continue

            info = annot.info or {}
            contents = info.get("content", "") or annot.get_text() or ""
            if not contents.strip() and annot_type == 2:
                # FreeText: try getting text from the annotation rect
                contents = annot.get_text() or ""

            # Split multi-comment annotations (e.g., "G1. xxx G2. yyy\rM1. zzz")
            split_contents = _split_comments(contents.strip())
            annot_type_str = ANNOT_TYPE_MAP.get(annot_type, f"Unknown({annot_type})")
            author = info.get("title", "")
            created_date = _parse_date(info.get("creationDate", ""))

            for single_comment in split_contents:
                comments.append({
                    "no": idx,
                    "drawing_no": drawing_no,
                    "page": page_num + 1,
                    "type": annot_type_str,
                    "author": author,
                    "contents": single_comment,
                    "reply": "",
                    "created_date": created_date,
                })
                idx += 1

    doc.close()

    return {
        "filename": filename,
        "drawing_no": drawing_no,
        "total_pages": total_pages,
        "total_comments": len(comments),
        "comments": comments,
    }


# ── Endpoints ──

@router.post("/extract")
async def extract_comments(file: UploadFile = File(...)):
    """Upload a PDF and extract all annotations/comments."""
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="PDF 파일만 업로드 가능합니다.")

    pdf_bytes = await file.read()
    return _extract_annotations(pdf_bytes, file.filename)


@router.post("/extract-blob")
async def extract_comments_from_blob(
    blob_path: str = Body(...),
    username: str = Body(None),
):
    """Azure Blob에서 PDF 다운로드 → 어노테이션 추출"""
    from app.services.blob_storage import get_container_client

    container_client = get_container_client()
    blob_client = container_client.get_blob_client(blob_path)

    if not blob_client.exists():
        raise HTTPException(status_code=404, detail=f"Blob not found: {blob_path}")

    pdf_bytes = blob_client.download_blob().readall()
    filename = blob_path.split('/')[-1]
    return _extract_annotations(pdf_bytes, filename)


@router.post("/delete-blob")
async def delete_comment_file(
    blob_path: str = Body(...),
    username: str = Body(None),
):
    """코멘트 PDF 삭제 (blob + JSON + Azure Search 인덱스)"""
    from app.services.blob_storage import get_container_client

    container_client = get_container_client()
    filename = blob_path.split('/')[-1]

    # 1. PDF blob 삭제
    blob = container_client.get_blob_client(blob_path)
    if blob.exists():
        blob.delete_blob()
        print(f"[Comments] Deleted PDF blob: {blob_path}", flush=True)

    # 2. JSON 분석 결과 삭제 ({username}/json/{base}.json)
    if username:
        base = os.path.splitext(filename)[0]
        json_blob = container_client.get_blob_client(f"{username}/json/{base}.json")
        if json_blob.exists():
            json_blob.delete_blob()
            print(f"[Comments] Deleted JSON: {username}/json/{base}.json", flush=True)

        # Also delete split-format JSON folder
        json_folder = f"{username}/json/{base}"
        try:
            split_blobs = list(container_client.list_blobs(name_starts_with=f"{json_folder}/"))
            for b in split_blobs:
                container_client.get_blob_client(b.name).delete_blob()
            if split_blobs:
                print(f"[Comments] Deleted {len(split_blobs)} split JSON blobs", flush=True)
        except Exception as e:
            print(f"[Comments] Warning: split JSON cleanup: {e}", flush=True)

    # 3. Azure Search 인덱스 삭제
    try:
        from app.services.azure_search import azure_search_service
        if azure_search_service.client:
            # Search by source filename and blob_path to scope deletion
            filter_expr = f"source eq '{filename}'"
            if username:
                filter_expr += f" and blob_path ge '{username}/' and blob_path lt '{username}0'"
            results = azure_search_service.client.search(
                search_text="*",
                filter=filter_expr,
                select=["id"],
                top=5000,
            )
            doc_ids = [{"id": r["id"]} for r in results]
            if doc_ids:
                azure_search_service.client.delete_documents(documents=doc_ids)
                print(f"[Comments] Deleted {len(doc_ids)} search index entries", flush=True)
    except Exception as e:
        print(f"[Comments] Warning: search index cleanup: {e}", flush=True)

    return {"deleted": blob_path}


@router.post("/export-excel")
async def export_excel(req: ExportRequest):
    """Convert table data to Excel and return as download."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl is not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PDF Comments"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["No", "도면번호", "페이지", "타입", "작성자", "코멘트 내용", "답변", "작성일자"]
    col_widths = [6, 25, 8, 12, 15, 50, 50, 18]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Data rows
    cell_align = Alignment(vertical="top", wrap_text=True)
    for row_idx, row in enumerate(req.rows, 2):
        values = [row.no, row.drawing_no, row.page, row.type, row.author, row.contents, row.reply, row.created_date]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = cell_align
            cell.border = thin_border

    # Freeze header row
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_filename = req.filename if req.filename.endswith(".xlsx") else req.filename + ".xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_filename}"'},
    )
