"""
Standalone PDF Annotation Extractor
====================================
Extract all annotations/comments from a PDF and export to Excel.

Usage:
    python extract_pdf_annotations.py <pdf_path> [output_path]

Example:
    python extract_pdf_annotations.py "../10-24000-OM-171-200_GENERAL ASSEMBLY DRAWING.pdf" comments.xlsx
"""

import re
import sys
from datetime import datetime
from pathlib import Path

import fitz  # PyMuPDF


ANNOT_TYPE_MAP = {
    0: "Text", 1: "Link", 2: "FreeText", 3: "Line", 4: "Square",
    5: "Circle", 6: "Polygon", 7: "PolyLine", 8: "Highlight",
    9: "Underline", 10: "Squiggly", 11: "StrikeOut", 12: "Stamp",
    13: "Caret", 14: "Ink", 15: "Popup", 16: "FileAttachment",
    17: "Sound", 18: "Movie", 19: "Widget", 20: "Screen",
    21: "PrinterMark", 22: "TrapNet", 23: "Watermark", 24: "3D", 25: "Redact",
}


def _extract_drawing_no(filename: str) -> str:
    name = filename.rsplit(".", 1)[0] if "." in filename else filename
    match = re.search(r'[\dA-Z][\dA-Z\-]{4,}[\dA-Z]', name, re.IGNORECASE)
    return match.group(0) if match else name


def _parse_date(date_str: str | None) -> str:
    if not date_str:
        return ""
    try:
        clean = date_str.replace("D:", "").split("+")[0].split("-")[0].split("Z")[0]
        if len(clean) >= 8:
            dt = datetime.strptime(clean[:14].ljust(14, "0"), "%Y%m%d%H%M%S")
            return dt.strftime("%Y-%m-%d %H:%M")
        return clean
    except Exception:
        return date_str


def extract_pdf_annotations(pdf_path: str) -> list[dict]:
    """Extract annotations from a PDF file. Returns list of comment dicts."""
    doc = fitz.open(pdf_path)
    filename = Path(pdf_path).name
    drawing_no = _extract_drawing_no(filename)
    comments = []
    idx = 1

    for page_num in range(len(doc)):
        page = doc[page_num]
        for annot in page.annots() or []:
            annot_type = annot.type[0] if annot.type else -1
            if annot_type in (1, 15):  # Skip Link/Popup
                continue

            info = annot.info or {}
            contents = info.get("content", "") or annot.get_text() or ""

            comments.append({
                "no": idx,
                "drawing_no": drawing_no,
                "page": page_num + 1,
                "type": ANNOT_TYPE_MAP.get(annot_type, f"Unknown({annot_type})"),
                "author": info.get("title", ""),
                "contents": contents.strip(),
                "reply": "",
                "created_date": _parse_date(info.get("creationDate", "")),
            })
            idx += 1

    doc.close()
    return comments


def extract_pdf_annotations_to_excel(pdf_path: str, output_path: str = None) -> str:
    """Extract annotations from PDF and save to Excel. Returns output path."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    comments = extract_pdf_annotations(pdf_path)

    if output_path is None:
        output_path = Path(pdf_path).stem + "_comments.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PDF Comments"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["No", "도면번호", "페이지", "타입", "작성자", "코멘트 내용", "답변", "작성일자"]
    col_widths = [6, 25, 8, 12, 15, 50, 50, 18]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    cell_align = Alignment(vertical="top", wrap_text=True)
    for row_idx, c in enumerate(comments, 2):
        values = [c["no"], c["drawing_no"], c["page"], c["type"], c["author"],
                  c["contents"], c["reply"], c["created_date"]]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = cell_align
            cell.border = thin_border

    ws.freeze_panes = "A2"
    wb.save(output_path)

    print(f"Extracted {len(comments)} annotations from '{pdf_path}'")
    print(f"Saved to: {output_path}")
    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_pdf_annotations.py <pdf_path> [output_path]")
        sys.exit(1)

    pdf_file = sys.argv[1]
    out_file = sys.argv[2] if len(sys.argv) > 2 else None
    extract_pdf_annotations_to_excel(pdf_file, out_file)
